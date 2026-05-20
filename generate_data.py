#!/usr/bin/env python3
"""
Run this script whenever the Faculty list Excel is updated.
Writes speakers.json (excluding Declined speakers) for the dashboard.

Usage:  python generate_data.py
"""
import json
import re
import unicodedata
import warnings
from pathlib import Path

import openpyxl

FACULTY_PATH = Path(r'C:\Users\carol\PycharmProjects\ICISA information\Faculty list follow up.xlsx')
PGM_PATH     = Path(r'C:\Users\carol\PycharmProjects\ICISA information\PGM ICISA 0305 (4).xlsx')
OUTPUT_PATH  = Path(__file__).parent / 'speakers.json'

EXCLUDE_TYPES = {'DO NOT INVITE'}


# ── helpers ─────────────────────────────────────────────────────────────────

def clean(v):
    return str(v).strip() if v is not None else ''


def norm(s):
    """Lowercase + strip diacritics for fuzzy matching."""
    s = str(s).lower().strip()
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def match_key(last_name):
    """
    Return the best single token to search for a speaker by last name.
    Handles hyphenated names (Shick-Nahtomi -> nahtomi) and
    multi-word last names (Ceyda Meco -> ceyda).
    """
    n = norm(last_name)
    parts = [p for p in re.split(r'[-\s]+', n) if len(p) > 3]
    if not parts:
        return n
    return max(parts, key=len)   # longest token = most distinctive


def get_status(row):
    s = norm(clean(row[20]) if len(row) > 20 else '')
    c = norm(clean(row[16]) if len(row) > 16 else '')
    if any(x in s for x in ('confirm', 'will come', 'unofficially', 'unofficialy')):
        return 'Confirmed'
    if any(x in s for x in ('declin', 'probably not', 'probaly', 'probably wont',
                              'probaby', 'wont come')):
        return 'Declined'
    if 'declin' in c:
        return 'Declined'
    return 'Pending'


# ── load speakers ────────────────────────────────────────────────────────────

def load_speakers():
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(FACULTY_PATH, read_only=True, data_only=True)

    ws = wb['International Speakers']
    speakers, seen = [], set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        first = clean(row[5])
        last  = clean(row[6])
        if not first or first == 'None':
            continue
        if 'moderator' in f'{first} {last}'.lower():
            continue

        stype = clean(row[3])
        if stype.upper() in EXCLUDE_TYPES:
            continue

        status = get_status(row)
        if status == 'Declined':
            continue                      # excluded per user request

        key = (norm(first), norm(last))
        if key in seen:
            continue
        seen.add(key)

        speakers.append({
            'first':  first,
            'last':   last,
            'type':   stype,
            'track':  clean(row[4]),
            'status': status,
            'bio':    'Yes' if (len(row) > 21 and row[21]) else '',
            'photo':  'Yes' if (len(row) > 22 and row[22]) else '',
            'tasks':  0,                  # filled below
        })

    wb.close()
    return speakers


# ── count tasks from program ─────────────────────────────────────────────────

def count_tasks(speakers):
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb2 = openpyxl.load_workbook(PGM_PATH, read_only=True, data_only=True)

    ws = wb2['Sheet1']
    time_re = re.compile(r'^\d{1,2}:\d{2}')
    mod_re  = re.compile(r'moderator', re.I)

    # Collect short lines (1-5 words) — these are where speaker names appear
    short_lines = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if text in ('None', ''):
                continue
            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                if time_re.match(line):
                    continue
                if mod_re.search(line):
                    continue
                if 1 <= len(line.split()) <= 5:
                    short_lines.append(norm(line))

    wb2.close()

    # Detect shared last-name keys so we can require first-name prefix for those
    from collections import Counter
    key_counts = Counter(match_key(s['last']) for s in speakers)

    for s in speakers:
        key = match_key(s['last'])
        if not key:
            continue
        key_re = re.compile(r'\b' + re.escape(key) + r'\b')

        if key_counts[key] > 1:
            # Multiple speakers share this last-name key → also require first-name prefix
            fp = norm(s['first'])[:4]
            fp_re = re.compile(r'\b' + re.escape(fp))
            s['tasks'] = sum(
                1 for line in short_lines
                if key_re.search(line) and fp_re.search(line)
            )
        else:
            s['tasks'] = sum(1 for line in short_lines if key_re.search(line))


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    speakers = load_speakers()
    count_tasks(speakers)

    total     = len(speakers)
    confirmed = sum(1 for s in speakers if s['status'] == 'Confirmed')
    pending   = sum(1 for s in speakers if s['status'] == 'Pending')
    no_tasks  = sum(1 for s in speakers if s['tasks'] == 0)
    heavy     = sum(1 for s in speakers if s['tasks'] >= 4)

    print(f'Speakers (excl. Declined): {total}')
    print(f'  Confirmed : {confirmed}')
    print(f'  Pending   : {pending}')
    print(f'  No tasks  : {no_tasks}')
    print(f'  4+ tasks  : {heavy}')
    print()
    print('Task counts per speaker:')
    for s in sorted(speakers, key=lambda x: -x['tasks']):
        print(f"  {s['tasks']:2d}  {s['first']} {s['last']}  [{s['status']}]")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(speakers, f, ensure_ascii=False, indent=2)
    print(f'\nWritten -> {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
